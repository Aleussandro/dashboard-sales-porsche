"""
Porsche Sales Data Sanitization Agent
=====================================
Reads base_dados_porsche.xlsx, applies all sanitization rules from schema.md,
and generates base_dados_porsche_tratado.xlsx in the material/ folder.
"""

import re
import datetime
import openpyxl
from pathlib import Path

# ============================================================================
# CONSTANTS & LOOKUP TABLES
# ============================================================================

CANONICAL_MODELS = [
    "911 Carrera GTS", "911 Carrera S", "911 Carrera",
    "911 Turbo S", "911 Turbo",
    "911 GT3 RS", "911 GT3",
    "911 Dakar",
    "911 Targa 4S", "911 Targa 4",
    "718 Cayman GT4 RS", "718 Cayman S", "718 Cayman",
    "718 Boxster GTS", "718 Boxster",
    "718 Spyder RS",
    "Cayenne E-Hybrid", "Cayenne Turbo GT", "Cayenne Turbo",
    "Cayenne Coupe", "Cayenne S", "Cayenne",
    "Macan Electric", "Macan GTS", "Macan S", "Macan T", "Macan",
    "Panamera 4 E-Hybrid", "Panamera Turbo S", "Panamera Turbo",
    "Panamera 4S", "Panamera 4", "Panamera",
    "Taycan Cross Turismo", "Taycan Turbo S", "Taycan Turbo",
    "Taycan GTS", "Taycan 4S", "Taycan",
]

PAYMENT_MAP = {
    "credit card": "Credit Card",
    "creditcard": "Credit Card",
    "credit": "Credit Card",
    "debit card": "Debit Card",
    "debitcard": "Debit Card",
    "debit": "Debit Card",
    "bank transfer": "Bank Transfer",
    "banktransfer": "Bank Transfer",
    "bank-transfer": "Bank Transfer",
    "bank wire": "Wire Transfer",
    "bankwire": "Wire Transfer",
    "wire transfer": "Wire Transfer",
    "wiretransfer": "Wire Transfer",
    "wire": "Wire Transfer",
    "financing": "Financing",
    "financing plan": "Financing",
    "financingplan": "Financing",
    "lease": "Lease",
    "leasing": "Lease",
    "cash": "Cash",
    "ach payment": "ACH Payment",
    "ach": "ACH Payment",
    "achpayment": "ACH Payment",
    "crypto payment": "Crypto Payment",
    "crypto": "Crypto Payment",
    "cryptopayment": "Crypto Payment",
    "cryptocurrency": "Crypto Payment",
}

STATE_MAP = {
    "alabama": "AL", "alaska": "AK", "arizona": "AZ", "arkansas": "AR",
    "california": "CA", "colorado": "CO", "connecticut": "CT", "delaware": "DE",
    "florida": "FL", "georgia": "GA", "hawaii": "HI", "idaho": "ID",
    "illinois": "IL", "indiana": "IN", "iowa": "IA", "kansas": "KS",
    "kentucky": "KY", "louisiana": "LA", "maine": "ME", "maryland": "MD",
    "massachusetts": "MA", "michigan": "MI", "minnesota": "MN",
    "mississippi": "MS", "missouri": "MO", "montana": "MT", "nebraska": "NE",
    "nevada": "NV", "new hampshire": "NH", "new jersey": "NJ",
    "new mexico": "NM", "new york": "NY", "north carolina": "NC",
    "north dakota": "ND", "ohio": "OH", "oklahoma": "OK", "oregon": "OR",
    "pennsylvania": "PA", "rhode island": "RI", "south carolina": "SC",
    "south dakota": "SD", "tennessee": "TN", "texas": "TX", "utah": "UT",
    "vermont": "VT", "virginia": "VA", "washington": "WA",
    "west virginia": "WV", "wisconsin": "WI", "wyoming": "WY",
    "district of columbia": "DC",
}
# Add reverse mapping (abbreviation -> abbreviation)
VALID_ABBREVS = set(STATE_MAP.values())
for abbr in list(VALID_ABBREVS):
    STATE_MAP[abbr.lower()] = abbr

DELIVERY_MAP = {
    "delivered": "Delivered",
    "deliverd": "Delivered",
    "pending": "Pending",
    "in transit": "In Transit",
    "in-transit": "In Transit",
    "intransit": "In Transit",
    "cancelled": "Cancelled",
    "canceled": "Cancelled",
    "awaiting delivery": "Awaiting Delivery",
    "awaitingdelivery": "Awaiting Delivery",
    "awaiting pickup": "Awaiting Pickup",
    "awaitingpickup": "Awaiting Pickup",
    "pending approval": "Pending Approval",
    "pendingapproval": "Pending Approval",
    "pending review": "Pending Review",
    "pendingreview": "Pending Review",
    "shipped": "Shipped",
    "awaiting review": "Awaiting Review",
    "awaitingreview": "Awaiting Review",
}

# Word-to-number mapping for text-based numbers
ONES = {
    "zero": 0, "one": 1, "two": 2, "three": 3, "four": 4, "five": 5,
    "six": 6, "seven": 7, "eight": 8, "nine": 9, "ten": 10,
    "eleven": 11, "twelve": 12, "thirteen": 13, "fourteen": 14, "fifteen": 15,
    "sixteen": 16, "seventeen": 17, "eighteen": 18, "nineteen": 19,
}
TENS = {
    "twenty": 20, "thirty": 30, "forty": 40, "fifty": 50,
    "sixty": 60, "seventy": 70, "eighty": 80, "ninety": 90,
}
SCALES = {
    "hundred": 100, "thousand": 1000, "million": 1000000, "billion": 1000000000,
}

MONTH_NAMES = {
    "january": 1, "february": 2, "march": 3, "april": 4,
    "may": 5, "june": 6, "july": 7, "august": 8,
    "september": 9, "october": 10, "november": 11, "december": 12,
    "jan": 1, "feb": 2, "mar": 3, "apr": 4,
    "jun": 6, "jul": 7, "aug": 8, "sep": 9, "sept": 9,
    "oct": 10, "nov": 11, "dec": 12,
}


# ============================================================================
# HELPER: WORDS TO NUMBER
# ============================================================================

def words_to_number(text: str) -> float | None:
    """Convert English number words to a numeric value.
    
    Handles compound forms like 'two hundred thousand', 'eighty two thousand',
    'twenty twenty four', 'twelve thousand', etc.
    """
    text = text.strip().lower()
    if not text:
        return None

    words = text.replace("-", " ").split()
    
    # Filter out non-number words
    number_words = set(ONES) | set(TENS) | set(SCALES)
    if not all(w in number_words for w in words):
        return None

    # Parse using standard English number word algorithm
    result = 0
    current = 0

    for word in words:
        if word in ONES:
            current += ONES[word]
        elif word in TENS:
            current += TENS[word]
        elif word == "hundred":
            if current == 0:
                current = 1
            current *= 100
        elif word == "thousand":
            if current == 0:
                current = 1
            current *= 1000
            result += current
            current = 0
        elif word == "million":
            if current == 0:
                current = 1
            current *= 1000000
            result += current
            current = 0
        elif word == "billion":
            if current == 0:
                current = 1
            current *= 1000000000
            result += current
            current = 0

    result += current
    return float(result) if result > 0 or text == "zero" else None


# ============================================================================
# SANITIZERS
# ============================================================================

def sanitize_date(raw) -> str:
    """Normalize sale_date to YYYY-MM-DD or INVALID."""
    if raw is None:
        return "INVALID"

    # Already a datetime object from openpyxl
    if isinstance(raw, datetime.datetime):
        return raw.strftime("%Y-%m-%d")
    if isinstance(raw, datetime.date):
        return raw.strftime("%Y-%m-%d")

    raw = str(raw).strip()
    if not raw:
        return "INVALID"

    # Try "Month DDth, YYYY" or "Month DD YYYY" patterns
    m = re.match(
        r'^(\w+)\s+(\d{1,2})(?:st|nd|rd|th)?,?\s*(\d{4})$', raw, re.IGNORECASE
    )
    if m:
        month_name, day_str, year_str = m.groups()
        month_num = MONTH_NAMES.get(month_name.lower())
        if month_num:
            try:
                d = datetime.date(int(year_str), month_num, int(day_str))
                return d.strftime("%Y-%m-%d")
            except ValueError:
                return "INVALID"

    # Try common delimiter-based formats
    for sep in ["-", "/", "."]:
        parts = raw.split(sep)
        if len(parts) == 3:
            p0, p1, p2 = [p.strip() for p in parts]
            
            # YYYY-MM-DD, YYYY/MM/DD, YYYY.MM.DD
            if len(p0) == 4 and p0.isdigit():
                try:
                    d = datetime.date(int(p0), int(p1), int(p2))
                    return d.strftime("%Y-%m-%d")
                except (ValueError, TypeError):
                    return "INVALID"
            
            # MM/DD/YYYY or MM-DD-YYYY
            if len(p2) == 4 and p2.isdigit():
                try:
                    d = datetime.date(int(p2), int(p0), int(p1))
                    return d.strftime("%Y-%m-%d")
                except (ValueError, TypeError):
                    return "INVALID"
            
            # MM/DD/YY or MM-DD-YY
            if len(p2) == 2 and p2.isdigit():
                year = 2000 + int(p2) if int(p2) < 50 else 1900 + int(p2)
                try:
                    d = datetime.date(year, int(p0), int(p1))
                    return d.strftime("%Y-%m-%d")
                except (ValueError, TypeError):
                    return "INVALID"

    return "INVALID"


def sanitize_model(raw) -> str:
    """Normalize porsche_model to canonical name or title-cased."""
    if raw is None:
        return "INVALID"
    raw = str(raw).strip()
    if not raw:
        return "INVALID"

    raw_lower = raw.lower()
    # Try exact match (case-insensitive) against canonical list
    # Sort canonical list by length descending so longer (more specific) matches win
    for canonical in CANONICAL_MODELS:
        if canonical.lower() == raw_lower:
            return canonical

    # Try if the raw input contains a canonical model name
    for canonical in CANONICAL_MODELS:
        if canonical.lower() in raw_lower:
            return canonical

    # Unknown model: title case
    return raw.title()


def sanitize_year(raw) -> str:
    """Normalize model_year to four-digit year or INVALID."""
    if raw is None:
        return "INVALID"

    raw_str = str(raw).strip()
    if not raw_str:
        return "INVALID"

    # Direct 4-digit integer
    if isinstance(raw, (int, float)):
        year = int(raw)
        if 1990 <= year <= 2035:
            return str(year)
        return "INVALID"

    # Try direct 4-digit string
    if re.match(r'^\d{4}$', raw_str):
        year = int(raw_str)
        if 1990 <= year <= 2035:
            return str(year)
        return "INVALID"

    # Format like "20-24" or "20 24"
    m = re.match(r'^(\d{2})[\s\-](\d{2})$', raw_str)
    if m:
        year = int(m.group(1) + m.group(2))
        if 1990 <= year <= 2035:
            return str(year)
        return "INVALID"

    # Text-based years
    # First try colloquial format: "twenty twenty four" -> 20|24 -> 2024
    # Split words and try to find a split point where left half gives a 2-digit
    # number (e.g. 19, 20) and right half gives a 2-digit number (00-99)
    text_lower = raw_str.lower().replace("-", " ").strip()
    text_words = text_lower.split()
    number_words = set(ONES) | set(TENS) | set(SCALES)
    if all(w in number_words for w in text_words) and "thousand" not in text_words:
        for split_idx in range(1, len(text_words)):
            left = " ".join(text_words[:split_idx])
            right = " ".join(text_words[split_idx:])
            left_num = words_to_number(left)
            right_num = words_to_number(right)
            if left_num is not None and right_num is not None:
                ln, rn = int(left_num), int(right_num)
                if 19 <= ln <= 20 and 0 <= rn <= 99:
                    year = ln * 100 + rn
                    if 1990 <= year <= 2035:
                        return str(year)

    # Standard text: "two thousand twenty two" -> 2022
    num = words_to_number(raw_str)
    if num is not None:
        year = int(num)
        if 1990 <= year <= 2035:
            return str(year)
        return "INVALID"

    return "INVALID"


def sanitize_price(raw) -> str:
    """Normalize sale_price to decimal USD amount with two decimals."""
    if raw is None:
        return "INVALID"

    if isinstance(raw, (int, float)):
        return f"{float(raw):.2f}"

    raw_str = str(raw).strip()
    if not raw_str:
        return "INVALID"

    text = raw_str.lower()

    # Remove common text/symbols
    text = text.replace("usd", "").replace("dollars", "").replace("dollar", "")
    text = text.replace("$", "").strip()

    # Check for word-based prices first: "eighty two thousand", etc.
    # Remove trailing/leading garbage
    text_words = re.sub(r'[^a-z\s\-]', '', text).strip()
    if text_words:
        num = words_to_number(text_words)
        if num is not None and num > 0:
            return f"{num:.2f}"

    # Now process numeric formats
    text = raw_str.lower()
    text = text.replace("usd", "").replace("dollars", "").replace("dollar", "")
    text = text.replace("$", "").strip()

    # Handle "k" suffix: "$645k" -> 645000
    m = re.match(r'^[\s]*([0-9.,]+)\s*k\s*$', text, re.IGNORECASE)
    if m:
        num_str = m.group(1).replace(",", "")
        try:
            return f"{float(num_str) * 1000:.2f}"
        except ValueError:
            pass

    # European-style: "89.750,00" (dots as thousand sep, comma as decimal)
    m = re.match(r'^[\s]*(\d{1,3}(?:\.\d{3})+),(\d{2})\s*$', text)
    if m:
        integer_part = m.group(1).replace(".", "")
        decimal_part = m.group(2)
        return f"{int(integer_part)}.{decimal_part}"

    # Format like "112.750" where dot is thousand separator (no decimal portion,
    # and integer part is 3+ digits with exactly 3 after dot)
    m = re.match(r'^[\s]*(\d{1,3})\.(\d{3})\s*$', text)
    if m:
        return f"{int(m.group(1) + m.group(2)):.2f}"

    # Standard: remove commas and parse
    cleaned = text.replace(",", "").strip()
    # Remove trailing non-numeric chars
    cleaned = re.sub(r'[^0-9.]', '', cleaned)
    if cleaned:
        try:
            val = float(cleaned)
            return f"{val:.2f}"
        except ValueError:
            pass

    return "INVALID"


def sanitize_mileage(raw) -> str:
    """Normalize vehicle_mileage to integer miles."""
    if raw is None:
        return "INVALID"

    if isinstance(raw, (int, float)):
        return str(int(raw))

    raw_str = str(raw).strip().lower()
    if not raw_str:
        return "INVALID"

    # Zero / new car
    if raw_str in ("new", "new car", "0", "zero", "zero miles", "0 mi", "0 miles"):
        return "0"
    if re.match(r'^(new|new\s+car|zero\s*(miles|mi\.?)?)$', raw_str, re.IGNORECASE):
        return "0"

    is_km = bool(re.search(r'\bkm\b', raw_str, re.IGNORECASE))

    # Word-based: "twelve thousand miles"
    text_only = re.sub(r'(miles?|mi\.?|km)', '', raw_str, flags=re.IGNORECASE).strip()
    text_only = re.sub(r'[^a-z\s\-]', '', text_only).strip()
    if text_only:
        num = words_to_number(text_only)
        if num is not None:
            miles = int(num)
            if is_km:
                miles = round(miles * 0.621371)
            return str(miles)

    # Extract numbers
    # Remove text labels
    cleaned = re.sub(r'(miles|mi\.?|km|:)', '', raw_str, flags=re.IGNORECASE).strip()

    # Handle European-style thousand separators with dots: "14.500"
    m = re.match(r'^[\s]*(\d{1,3})\.(\d{3})\s*$', cleaned)
    if m:
        value = int(m.group(1) + m.group(2))
        if is_km:
            value = round(value * 0.621371)
        return str(value)

    # Handle "KM 18,900" style (comma as thousand separator)
    cleaned = cleaned.replace(",", "").strip()
    # Remove any remaining non-numeric characters except dot
    cleaned = re.sub(r'[^0-9.]', '', cleaned)
    if cleaned:
        try:
            value = int(float(cleaned))
            if is_km:
                value = round(value * 0.621371)
            return str(value)
        except ValueError:
            pass

    return "INVALID"


def sanitize_payment(raw) -> str:
    """Normalize payment_method to controlled label."""
    if raw is None:
        return "INVALID"
    raw_str = str(raw).strip()
    if not raw_str:
        return "INVALID"

    # Clean: lowercase, remove hyphens, extra spaces
    key = raw_str.lower().replace("-", " ").replace("_", " ")
    key = re.sub(r'\s+', ' ', key).strip()

    if key in PAYMENT_MAP:
        return PAYMENT_MAP[key]

    # Try without spaces
    key_nospace = key.replace(" ", "")
    if key_nospace in PAYMENT_MAP:
        return PAYMENT_MAP[key_nospace]

    # Unknown: title case
    return raw_str.title()


def sanitize_city(raw) -> str:
    """Normalize city to title case, preserving known punctuation."""
    if raw is None:
        return "INVALID"
    raw_str = str(raw).strip()
    if not raw_str:
        return "INVALID"

    # Title case
    result = raw_str.title()
    
    # Fix common abbreviation patterns broken by title()
    # "St." should stay as "St."
    result = re.sub(r'\bSt\b(?!\.)', 'St.', result)
    
    return result


def sanitize_state(raw) -> str:
    """Normalize US state to USPS two-letter code or INVALID."""
    if raw is None:
        return "INVALID"
    raw_str = str(raw).strip()
    if not raw_str:
        return "INVALID"

    key = raw_str.lower().strip()
    if key in STATE_MAP:
        return STATE_MAP[key]

    return "INVALID"


def sanitize_delivery(raw) -> str:
    """Normalize delivery_status to controlled label."""
    if raw is None:
        return "INVALID"
    raw_str = str(raw).strip()
    if not raw_str:
        return "INVALID"

    # Clean: remove punctuation (!!!, .), lowercase, normalize whitespace/hyphens
    cleaned = re.sub(r'[!.,;:]+', '', raw_str)
    cleaned = cleaned.lower().replace("-", " ").replace("_", " ")
    cleaned = re.sub(r'\s+', ' ', cleaned).strip()

    if cleaned in DELIVERY_MAP:
        return DELIVERY_MAP[cleaned]

    # Try without spaces
    nospace = cleaned.replace(" ", "")
    if nospace in DELIVERY_MAP:
        return DELIVERY_MAP[nospace]

    # Unknown: title case
    return raw_str.title()


# ============================================================================
# MAIN PROCESSING
# ============================================================================

def process(input_path: str, output_path: str):
    """Read the input XLSX, sanitize all columns, write the output XLSX."""
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active

    # Read header row
    headers = [cell.value for cell in ws[1]]
    
    # Define the sanitization map: (source_col_name, output_col_name, sanitizer_fn)
    sanitizers = [
        ("sale_date",        "SaleDateSanitized",       sanitize_date),
        ("porsche_model",    "PorscheModelSanitized",    sanitize_model),
        ("model_year",       "ModelYearSanitized",       sanitize_year),
        ("sale_price",       "SalesPriceSanitized",      sanitize_price),
        ("vehicle_mileage",  "VehicleMileageSanitized",  sanitize_mileage),
        ("payment_method",   "PayMethodSanitized",       sanitize_payment),
        ("city",             "CitySanitized",            sanitize_city),
        ("state",            "StateSanitized",           sanitize_state),
        ("delivery_status",  "DeliveryStatusSanitized",  sanitize_delivery),
    ]

    # Build source column index lookup
    col_index = {name: i for i, name in enumerate(headers)}

    # Read all data rows
    data_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data_rows.append(list(row))

    # Build new headers with sanitized columns inserted after their source
    # Work from right to left so insertions don't shift indices
    new_headers = list(headers)
    insert_plan = []  # (insert_position, sanitizer_index)
    
    for san_src, san_out, san_fn in sanitizers:
        src_idx = col_index[san_src]
        insert_plan.append((src_idx, san_out, san_fn, san_src))

    # Sort by source index descending so we can insert right-to-left
    insert_plan.sort(key=lambda x: x[0], reverse=True)

    for src_idx, san_out, san_fn, san_src in insert_plan:
        insert_pos = src_idx + 1
        new_headers.insert(insert_pos, san_out)
        for row in data_rows:
            raw_value = row[src_idx]
            sanitized = san_fn(raw_value)
            row.insert(insert_pos, sanitized)

    # Create output workbook
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Sanitized Data"

    # Write headers
    for col_num, header in enumerate(new_headers, 1):
        cell = ws_out.cell(row=1, column=col_num, value=header)
        cell.font = openpyxl.styles.Font(bold=True)

    # Write data
    for row_num, row_data in enumerate(data_rows, 2):
        for col_num, value in enumerate(row_data, 1):
            ws_out.cell(row=row_num, column=col_num, value=value)

    # Auto-adjust column widths
    for col in ws_out.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws_out.column_dimensions[col_letter].width = min(max_length + 3, 40)

    wb_out.save(output_path)
    print(f"✅ Arquivo sanitizado salvo em: {output_path}")
    print(f"   Linhas de dados: {len(data_rows)}")
    print(f"   Colunas: {len(new_headers)} ({len(headers)} originais + {len(sanitizers)} sanitizadas)")

    return new_headers, data_rows


# ============================================================================
# ENTRY POINT
# ============================================================================

if __name__ == "__main__":
    import sys

    if len(sys.argv) >= 2:
        input_file = Path(sys.argv[1])
        if len(sys.argv) >= 3:
            output_file = Path(sys.argv[2])
        else:
            # Auto-generate output name: arquivo.xlsx -> arquivo_tratado.xlsx
            output_file = input_file.parent / f"{input_file.stem}_tratado.xlsx"
    else:
        # Default: original Porsche file
        base_dir = Path(__file__).parent / "material"
        input_file = base_dir / "base_dados_porsche.xlsx"
        output_file = base_dir / "base_dados_porsche_tratado.xlsx"

    print(f"📂 Entrada:  {input_file}")
    print(f"📂 Saída:    {output_file}")
    print()

    headers, data = process(str(input_file), str(output_file))

    # Print sample of first 5 rows for verification
    print("\n📋 Amostra das primeiras 5 linhas (colunas sanitizadas):")
    san_cols = [i for i, h in enumerate(headers) if "Sanitized" in str(h)]
    san_headers = [headers[i] for i in san_cols]
    print(f"   {san_headers}")
    for row in data[:5]:
        print(f"   {[row[i] for i in san_cols]}")

